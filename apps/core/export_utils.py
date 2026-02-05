"""
Export Utilities - PDF va Excel export funksiyalari.
Barcha ilovalarda foydalanish mumkin.
"""
import io
from datetime import datetime
from decimal import Decimal
from django.http import HttpResponse
from django.utils import timezone

# Excel
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# PDF
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


# ========================================
# EXCEL EXPORT
# ========================================

def export_to_excel(data, columns, filename, title=None, sheet_name="Ma'lumotlar"):
    """
    Universal Excel export funksiyasi.

    Args:
        data: List of dicts yoki list of lists
        columns: [{'key': 'field_name', 'header': 'Sarlavha', 'width': 20}, ...]
        filename: Fayl nomi (`.xlsx` qo'shiladi)
        title: Sarlavha (ixtiyoriy)
        sheet_name: Sheet nomi

    Returns:
        HttpResponse with Excel file
    """
    if not EXCEL_AVAILABLE:
        return HttpResponse("Excel kutubxonasi o'rnatilmagan", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    number_alignment = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row_num = 1

    # Title (agar berilgan bo'lsa)
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        row_num = 3

    # Headers
    for col_num, column in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num, value=column.get('header', column.get('key', '')))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

        # Column width
        width = column.get('width', 15)
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Data rows
    for row_data in data:
        row_num += 1
        for col_num, column in enumerate(columns, 1):
            key = column.get('key')

            # Dict yoki list/tuple
            if isinstance(row_data, dict):
                value = row_data.get(key, '')
            elif isinstance(row_data, (list, tuple)):
                value = row_data[col_num - 1] if col_num - 1 < len(row_data) else ''
            else:
                value = getattr(row_data, key, '') if hasattr(row_data, key) else ''

            # Callable bo'lsa
            if callable(value):
                value = value()

            # None -> ''
            if value is None:
                value = ''

            # Decimal -> float
            if isinstance(value, Decimal):
                value = float(value)

            # DateTime format
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M')

            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border

            # Number alignment
            if isinstance(value, (int, float)):
                cell.alignment = number_alignment
                # Money format
                if column.get('money'):
                    cell.number_format = '#,##0'
            else:
                cell.alignment = cell_alignment

    # Freeze header
    ws.freeze_panes = f'A{3 if title else 2}'

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'

    wb.save(response)
    return response


# ========================================
# PDF EXPORT
# ========================================

def export_to_pdf(data, columns, filename, title=None, subtitle=None, landscape_mode=False):
    """
    Universal PDF export funksiyasi.

    Args:
        data: List of dicts yoki list of lists
        columns: [{'key': 'field_name', 'header': 'Sarlavha', 'width': 2*cm}, ...]
        filename: Fayl nomi (`.pdf` qo'shiladi)
        title: Sarlavha
        subtitle: Qo'shimcha ma'lumot
        landscape_mode: Horizontal format

    Returns:
        HttpResponse with PDF file
    """
    if not PDF_AVAILABLE:
        return HttpResponse("PDF kutubxonasi o'rnatilmagan", status=500)

    buffer = io.BytesIO()

    # Page size
    page_size = landscape(A4) if landscape_mode else A4

    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1*cm,
        bottomMargin=1*cm
    )

    elements = []
    styles = getSampleStyleSheet()

    # Title
    if title:
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=10,
            alignment=1  # Center
        )
        elements.append(Paragraph(title, title_style))

    # Subtitle
    if subtitle:
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=20,
            alignment=1,
            textColor=colors.gray
        )
        elements.append(Paragraph(subtitle, subtitle_style))

    elements.append(Spacer(1, 0.5*cm))

    # Table data
    table_data = []

    # Headers
    headers = [col.get('header', col.get('key', '')) for col in columns]
    table_data.append(headers)

    # Data rows
    for row_data in data:
        row = []
        for column in columns:
            key = column.get('key')

            if isinstance(row_data, dict):
                value = row_data.get(key, '')
            elif isinstance(row_data, (list, tuple)):
                idx = columns.index(column)
                value = row_data[idx] if idx < len(row_data) else ''
            else:
                value = getattr(row_data, key, '') if hasattr(row_data, key) else ''

            if callable(value):
                value = value()

            if value is None:
                value = ''

            if isinstance(value, Decimal):
                value = f"{value:,.0f}"

            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M')

            # Money format
            if column.get('money') and isinstance(value, (int, float)):
                value = f"{value:,.0f}"

            row.append(str(value))
        table_data.append(row)

    # Column widths
    col_widths = [col.get('width', 3*cm) for col in columns]

    # Create table
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Table style
    style = TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),

        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),

        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),

        # Padding
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ])

    table.setStyle(style)
    elements.append(table)

    # Footer
    elements.append(Spacer(1, 1*cm))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.gray,
        alignment=2  # Right
    )
    footer_text = f"Yaratilgan: {timezone.now().strftime('%Y-%m-%d %H:%M')} | SMART EDU CRM"
    elements.append(Paragraph(footer_text, footer_style))

    # Build PDF
    doc.build(elements)

    # Response
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'

    return response


# ========================================
# HELPER FUNCTIONS
# ========================================

def format_money(value):
    """Pul formatini qaytaradi: 1,500,000"""
    if value is None:
        return "0"
    return f"{value:,.0f}"


def format_date(value, fmt='%Y-%m-%d'):
    """Sana formatini qaytaradi"""
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return value.strftime(fmt)


def format_datetime(value, fmt='%Y-%m-%d %H:%M'):
    """Sana-vaqt formatini qaytaradi"""
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return value.strftime(fmt)
